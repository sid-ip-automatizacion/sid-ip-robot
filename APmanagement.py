import re
import openpyxl
import requests
import tkinter
import time
from pathlib import Path
from tkinter import ttk
from tkinter import filedialog

from AccessPoint import AccessPoint


def error_window(text):
        print("ERROR: {}".format(text))
        error_win = tkinter.Tk()
        error_text = tkinter.Label(error_win, text=text)
        ok_button = tkinter.Button(error_win, text='OK', width=5, height=2, command=error_win.destroy)
        error_text.pack()
        ok_button.pack()
        error_win.mainloop()

class ControllerSpecific():
    """
    Mantiene información de la controladora
    """
    def __init__(self, meraki_key='NA'):
        self.vendor = None
        self.login_user = None             # Usuario
        self.login_pass = None             # Contraseña
        self.ip = None                     # ip de la controladora
        self.sz_version = 'v9_1'           # smartzone API version
        self.sz_url = None                 # Inicializa la url
        self.meraki_api_key = meraki_key   # Key de meraki
        self.forti_key = None              # Fotigate API key
        self.forti_vdom = None             # Fortigate vdom

    def set_ip(self, ip):
        ip_valid = False
        ip_pattern = re.compile(
            '^(25[0-5]\.|2[0-4][0-9]\.|1[0-9][0-9]\.|[0-9][0-9]\.|[0-9]\.){3}(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9][0-9]|[0-9])(:[0-9]{0,4})?$')
        if ip_pattern.fullmatch(ip) != None:
            ip_valid = True
        elif self.vendor == 'meraki':
            print("meraki doesnt require ip")
            ip_valid = True
        else:
            error_window("Verify the controller IP")
            raise Exception("invalid ip")

        if (self.vendor == 'ruckus_vsz' or self.vendor == 'ruckus_sz_onsite') and ip_valid:
            self.ip = ip
            self.sz_url = 'https://{}/wsg/api/public/'.format(self.ip)

        if self.vendor == 'fortinet':
            self.ip = ip

    def set_sz_api_version(self,isold):
        if isold:
            self.sz_version = 'v8_0'
        else:
            self.szversion = 'v9_1'

class SZ_Communication():
    def __init__(self, sz_url, sz_version, pass_sz, user_sz):
        self.session = requests.Session()
        myheaders = {
            'Content-Type': 'application/json',
        }
        print('{SZurl}{ver}/serviceTicket'.format(SZurl=sz_url, ver=sz_version))
        try:
            res_serviceTicket = self.session.post('{SZurl}{ver}/serviceTicket'.format(SZurl=sz_url, ver=sz_version),
                                                  headers=myheaders, json={'username': user_sz, 'password': pass_sz}, verify=False, timeout=5)
            print("Abrir service ticket:", res_serviceTicket)
            self.serviceTicket = res_serviceTicket.json()['serviceTicket']
            print(self.serviceTicket)
        except:
            print("error obteniendo service ticket")
            error_window("Problem connecting to controller")

    def close_sz_session(self, sz_url, sz_version):
        param_st = {'serviceTicket': self.serviceTicket}
        self.session.delete('{SZurl}{ver}'.format(SZurl=sz_url, ver=sz_version), params=param_st)
        self.session.close()


class MerakiCommunicator():
    def __init__(self, api_key):
        self.myheaders = {
            'x-cisco-meraki-api-key': api_key
        }

    def send_request(self, url):
        session = requests.Session()
        try:
            request = session.get(url, headers=self.myheaders)
        except:
            print("error conectandose a meraki")
            error_window("Problem connecting to meraki portal")
        return request.json()

    def put_requests(self, url, r_body):
        try:
            response = requests.request('PUT', url, headers=self.myheaders, data=r_body)
        except:
            print("error conectandose a meraki")
            error_window("Problem connecting to meraki portal")
        print(response)


def main_function(root_win, meraki_key_api):
    controller = ControllerSpecific(meraki_key=meraki_key_api)

    def create_aplist_from_excel(vendor, conf_file):
        """
        Crea la lista de APs a partir del archivo de excel
        :param vendor: vendor de los APs (Meraki, Ruckus_vsz)
        :param conf_file: ruta al archivo con los nombres, descripciones, seriales y MACs de los APs
        :return: Lista de objetos tipo AccessPoint creados a partir del archivo de excel
        """
        wb = openpyxl.load_workbook(conf_file)
        ap_sh = wb['APinfo']
        new_ap_list = []
        for ap_row in ap_sh.iter_rows(min_row=2):
            if ap_row[0].value != None:
                new_ap = AccessPoint(vendor)
                new_ap.name = ap_row[0].value.strip()
                new_ap.description = ap_row[2].value.strip()
                new_ap.mac = ap_row[5].value.upper().strip().replace('-',':')
                new_ap.serial = ap_row[6].value.upper().strip()
                new_ap.site = ap_row[3].value.upper().strip()
                if vendor == 'ruckus_vsz' or vendor == 'ruckus_sz_onsite':
                    new_ap.controller_ip = controller.ip
                new_ap_list.append(new_ap)
        return new_ap_list


    def label_aps(vendor, filepath):
        """
        Configura en la controladora los nombres y descripciones de los APs especificados en el archivo de excel
        """
        conf_list = create_aplist_from_excel(vendor, filepath)
        if vendor == 'ruckus_vsz':
            if controller.login_user and controller.login_pass:
                vSZ_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                for ap in conf_list:
                    ap.config_ap(SZserviceTicket=vSZ_comm.serviceTicket, SZapiversion=controller.sz_version)
                    time.sleep(0.21)
                vSZ_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("password or user not defined")
        elif vendor == 'ruckus_sz_onsite':
            if controller.login_user and controller.login_pass:
                sz_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                for ap in conf_list:
                    ap.config_ap(SZserviceTicket=sz_comm.serviceTicket, SZapiversion=controller.sz_version)
                    time.sleep(0.21)
                sz_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("password or user not defined")
        elif vendor == 'meraki':
            for ap in conf_list:
                ap.config_ap(meraki_api_key=controller.meraki_api_key)
                time.sleep(0.21)
        elif vendor == 'fortinet':
            if controller.forti_key:
                for ap in conf_list:
                    ap.config_ap(forti_api_key=controller.forti_key, forti_IP=controller.ip)
                    time.sleep(0.21)
            else:
                error_window("You must define Fortigate API Key")
                raise Exception("API key not defined")
        else:
            print("Vendor no encontrado")


    def select_client(vendor):
        """
        Selecciona un cliente específico de la lista actual de clientes de la controladora
        :param vendor: vendor de los APs (Meraki, Ruckus_vsz)
        :return: ID del cliente seleccionado, reconocible por la controladora
        """
        clients = [] # listado con tuplas del nombre y id de los dominios
        skip_selection = False
        if vendor == 'ruckus_vsz':
            if controller.login_user and controller.login_pass:
                vSZ_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                param_st = {'serviceTicket': vSZ_comm.serviceTicket, 'listSize': 500}
                res_domains = vSZ_comm.session.get('{vSZurl}{ver}/domains'.format(vSZurl=controller.sz_url, ver=controller.sz_version),
                                                   params=param_st)
                list_domains = res_domains.json()['list'] # Lista completa con información de todos los dominios
                for count, domain in enumerate(list_domains):
                    clientinfo = (domain['name'], domain['id'])
                    clients.append(clientinfo)
                    print('{}. {} has id: {}'.format(count, domain['name'], domain['id']))
                vSZ_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("password or user not defined")
        elif vendor == 'ruckus_sz_onsite':
            print("client selection not required for smartzone onsite")
            skip_selection = True
        elif vendor == 'meraki':
            meraki_comm = MerakiCommunicator(controller.meraki_api_key)
            array_orgs = meraki_comm.send_request('https://api.meraki.com/api/v1/organizations')
            for count, organization in enumerate(array_orgs):
                clientinfo = (organization['name'], organization['id'])
                clients.append(clientinfo)
                print('{}. {} has id: {}'.format(count, organization['name'], organization['id']))
        elif vendor == 'fortinet':
            print("client selection not required for fortigate onsite")
            skip_selection = True
        else:
            print("error: vendor desconocido")

        clients.sort()

        class ClientSelected:
            def __init__(self):
                self.name = ''
                self.clientID = ''

            def select_client(self, name, id_client):
                self.name = name
                self.clientID = id_client
                select_win.quit()

        list_button_clients = []
        def search_client(event):
            cadena = search_text.get("1.0", "end").replace("\n", "").lower()
            for i in range(len(list_button_clients)):
                if len(cadena) >= 3 and cadena in list_button_clients[i]["text"].lower():
                    list_button_clients[i].config(bg='yellow')
                else:
                    list_button_clients[i].config(bg='white smoke')

        if not skip_selection:
            # Dibuja el listado para elegir
            def frame_configure(event):
                my_canvas.config(width=event.width, height=event.height)
                select_win.update_idletasks()
                my_canvas.configure(scrollregion=my_canvas.bbox(windows_item))
            select_win = tkinter.Tk()
            select_win.geometry('800x600')
            select_win.title('Select client')
            main_frame = tkinter.Frame(select_win)
            main_frame.pack(fill=tkinter.BOTH, expand=True)
            main_frame.rowconfigure(0, weight=1, minsize=10)
            main_frame.rowconfigure(1, weight=1, minsize=5)
            main_frame.columnconfigure(0, weight=1, minsize=10)
            main_frame.columnconfigure(1, weight=1, minsize=5)
            my_canvas = tkinter.Canvas(main_frame)
            scrollbar_vertical = ttk.Scrollbar(main_frame, orient='vertical', command=my_canvas.yview)
            scrollbar_horizontal = ttk.Scrollbar(main_frame, orient='horizontal', command=my_canvas.xview)
            work_area = tkinter.Frame(master=my_canvas, padx=10, pady=10)
            windows_item = my_canvas.create_window((0, 0), window=work_area, anchor='nw')
            main_frame.bind('<Configure>', frame_configure)
            my_canvas.configure(yscrollcommand=scrollbar_vertical.set, xscrollcommand=scrollbar_horizontal.set)
            my_canvas.grid(row=0, column=0, sticky='nsew')
            scrollbar_horizontal.grid(row=1, column=0, sticky="ew", ipadx=10, ipady=10)
            scrollbar_vertical.grid(row=0, column=1, sticky="ns", ipadx=10, ipady=10)

            client = ClientSelected()
            work_area.columnconfigure(0, weight=1, minsize=40)
            work_area.columnconfigure(1, weight=1, minsize=10)
            work_area.rowconfigure(0, weight=1, minsize=2)
            work_area.rowconfigure(1, weight=1, minsize=10)
            sel_frame_right = tkinter.Frame(master=work_area)
            sel_frame_left = tkinter.Frame(master=work_area, bg='white smoke')
            sel_frame_right.grid(row=0, column=1, padx=(3, 0), sticky="n")
            sel_frame_left.grid(row=0, column=0, sticky="nsew")

            sel_frame_left.rowconfigure(0, weight=1, minsize=10)
            sel_frame_left.columnconfigure(0, weight=1, minsize=20)

            for count, clint in enumerate(clients):
                select_win.rowconfigure(count, weight=1, minsize=1)
                but_client_name = tkinter.Button(master=sel_frame_left, text=clint[0], height=1,
                                                 command=lambda name=clint[0], id_client=clint[1]: client.select_client(name,id_client))
                but_client_name.grid(row=count, column=0, sticky="ew")
                list_button_clients.append(but_client_name)

            sel_frame_right.columnconfigure(0, weight=1, minsize=10)
            sel_frame_right.columnconfigure(1, weight=1, minsize=10)
            sel_frame_right.rowconfigure(0, weight=1, minsize=10)

            # search_text
            searchlabel = tkinter.Label(master=sel_frame_right, text="Search: ")
            searchlabel.grid(row=0, column=0, sticky="n")
            search_text = tkinter.Text(master=sel_frame_right, width=15, height=1)
            search_text.grid(row=0, column=1, sticky="n")
            search_text.bind('<KeyRelease>', search_client)

            select_win.mainloop()
            select_win.destroy()
        else :
            client = ClientSelected()
            client.ClientID = 'na'
        return client.clientID


    def get_aps_client(client_id, vendor):
        """
        Lee en la controladora una lista de los APs existentes en el cliente
        :param client_id: ID del cliente reconocible por la controladora
        :param vendor: vendor de los APs (Meraki, Ruckus_vsz)
        :return: Lista con objetos tipo AccessPoint creados de los APs existentes en la controladora
        """
        read_ap_client =[]
        if vendor == 'ruckus_vsz':
            if controller.login_user and controller.login_pass:
                vSZ_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                param_st = {'serviceTicket': vSZ_comm.serviceTicket, 'domainId': client_id, 'listSize': 999}
                res_aps_domain = vSZ_comm.session.get('{vSZurl}{ver}/aps'.format(vSZurl=controller.sz_url, ver=controller.sz_version),
                                                      params=param_st)
                list_aps = res_aps_domain.json()['list']  # Lista completa de APs del dominio
                for ap in list_aps:
                    apinfo = (ap['mac'], ap['serial'])
                    read_ap_client.append(apinfo)
                print("lee todos los aps del cliente {} de la vSZ".format(client_id))
                vSZ_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("password or user not defined")
        elif vendor == 'ruckus_sz_onsite':
            if controller.login_user and controller.login_pass:
                sz_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                param_st = {'serviceTicket': sz_comm.serviceTicket, 'listSize': 999}
                res_aps = sz_comm.session.get('{SZurl}{ver}/aps'.format(SZurl=controller.sz_url, ver=controller.sz_version),
                                              params=param_st)
                list_aps = res_aps.json()['list'] # Lista completa de APs de la controladora
                for ap in list_aps:
                    apinfo = (ap['mac'], ap['serial'])
                    read_ap_client.append(apinfo)
                print("lee todos los aps de la controladora")
                sz_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("password or user not defined")
        elif vendor == 'meraki':
            meraki_comm = MerakiCommunicator(controller.meraki_api_key)
            devices = meraki_comm.send_request('https://api.meraki.com/api/v1/organizations/{}/devices'.format(client_id))
            mr_pattern = re.compile('^MR\w+', re.IGNORECASE)
            for ap in devices:
                if mr_pattern.fullmatch(ap['model']) != None:
                    apinfo = (ap['mac'], ap['serial'])
                    read_ap_client.append(apinfo)
                print("El equipo {} no es un modelo MR".format(ap['serial']))
            print("lee todos los aps del cliente {} del portal meraki".format(client_id))
        elif vendor == 'fortinet':
            if controller.forti_key:
                url = "https://{fIP}/api/v2/monitor/wifi/managed_ap?" \
                      "vdom=*&access_token={key}".format(fIP=controller.ip, key=controller.forti_key)
                print(url)
                requests.packages.urllib3.disable_warnings()
                data = requests.get(url, verify=False)
                print("request response code : .............................. ", data.status_code)
                main_dict = data.json()
                main_dic_results = []
                for vdom_ans in main_dict:
                    if len(vdom_ans['results']) > 0:
                        main_dic_results.extend(vdom_ans['results'])
                for ap in range(len(main_dic_results)):
                    apinfo = (main_dic_results[ap]['board_mac'], main_dic_results[ap]['serial'])
                    read_ap_client.append(apinfo)
                print("lee todos los aps del fortigate")
            else:
                error_window("You must fortigate key")
                raise Exception("fortigate not defined")
        else:
            print("error: vendor desconocido")
        new_ap_list = []
        for ap in range(len(read_ap_client)):
            new_ap_list.append(AccessPoint(vendor))
            new_ap_list[ap].mac = read_ap_client[ap][0]
            new_ap_list[ap].serial = read_ap_client[ap][1]
            if vendor == 'ruckus_vsz' or vendor == 'ruckus_sz_onsite' or vendor=='fortinet':
                new_ap_list[ap].controller_ip = controller.ip
        return new_ap_list

    def complete_mac_or_serial(vendor, file_path, new_file):
        """
        Completa las MAC o Seriales de la tabla de excel
        :param vendor: vendor de los APs (Meraki, Ruckus_vsz)
        :param file_path: ruta al excel de APs con la informacion a completar
        :param new_file: ruta al excel donde se guardará el archivo completado
        """
        file_pass = validate_format(file_path)
        if file_pass:
            client_id = select_client(vendor)
            print("ID de cliente", client_id)
            aps_client = get_aps_client(client_id, vendor)
            ap_wb = openpyxl.load_workbook(file_path)
            ap_sh = ap_wb['APinfo']
            # Revisa todas las lineas del archivo de excel
            for ap_row in ap_sh.iter_rows(min_row=2):
                # Se tiene la MAC pero no el Serial
                if ap_row[5].value != None and ap_row[6].value == None:
                    ap_mac_excel = ap_row[5].value.upper().strip().replace('-', ':')
                    # Busca un ap de la lista por MAC y copia su serial en excel
                    mac_found = False
                    for ap in aps_client:
                        if ap.mac.upper() == ap_mac_excel:
                            ap_row[6].value = ap.serial
                            mac_found = True
                            break
                    if not mac_found:
                        print("El AP con MAC address {} no se encontró en el cliente".format(ap_mac_excel))
                # Se tiene el Serial pero no la MAC
                elif ap_row[5].value == None and ap_row[6].value != None:
                    ap_serial_excel = ap_row[6].value.upper().strip()
                    # Busca un ap de la lista por Serial, y copia su MAC en excel
                    serial_found = False
                    for ap in aps_client:
                        if ap.serial.upper() == ap_serial_excel:
                            ap_row[5].value = ap.mac
                            serial_found = True
                            break
                    if not serial_found:
                        print("El AP con Serial {} no se encontró en el cliente".format(ap_serial_excel))
                # Se tiene serial y MAC
                elif ap_row[5].value != None and ap_row[6].value != None:
                    print("El equipo {} está completo".format(ap_row[0].value))
                # No se tiene serial ni MAC
                elif ap_row[5].value == None and ap_row[6].value == None and ap_row[0].value != None:
                    print("El equipo {} no tiene MAC ni serial. Se debe tener MAC o serial de todos los equipos".format(ap_row[0].value))
                else:
                    print("Equipo {} no requiere completarse".format(ap_row[0].value))
            ap_wb.save(new_file)
        else:
            print("Errores en los datos suministrados")

    def validate_format(file_path):
        """
        Valida si los campos de un archivo de configuración tienen formato válido
        :param file_path: Archivo de configuración a validar
        :return: False si formato no es válido
        """
        format_valid = False
        mac_pattern = re.compile('((\d|[a-f]){2}:){5}(\d|[a-f]){2}',re.IGNORECASE)
        serial_pattern = re.compile('^\w[\w-]*\w$',re.IGNORECASE)
        ip_pattern = re.compile('^(25[0-5]\.|2[0-4][0-9]\.|1[0-9][0-9]\.|[0-9][0-9]\.|[0-9]\.){3}(25[0-5]|2[0-4][0-9]|1[0-9][0-9]|[1-9][0-9]|[0-9])$')
        name_pattern = re.compile('.{1,64}')
        descr_pattern = re.compile('.+')
        ap_wb = openpyxl.load_workbook(file_path)
        ap_sh = ap_wb['APinfo']
        # Revisa todas las lineas del archivo de excel
        for count, ap_row in enumerate(ap_sh.iter_rows(min_row=2), start=1):
            if ap_row[0].value != None:
                if name_pattern.fullmatch(ap_row[0].value) != None:
                    format_valid = True
                    print("Fila {}: AP name sin errores".format(count))
                else:
                    print("Fila {}: ERROR en AP name".format(count))
                    format_valid = False
            else:
                print("Fila {}: AP name vacio".format(count))
            if ap_row[2].value != None:
                if descr_pattern.fullmatch(ap_row[2].value) != None:
                    format_valid = True
                    print("Fila {}: Description sin errores".format(count))
                else:
                    format_valid = False
                    print("Fila {}: ERROR en Description".format(count))
            else:
                print("Fila {}: Description vacio".format(count))
            if ap_row[4].value != None:
                if ip_pattern.fullmatch(ap_row[4].value) != None:
                    format_valid = True
                    print("Fila {}: IP sin errores".format(count))
                else:
                    format_valid = False
                    print("Fila {}: ERROR en IP".format(count))
            else:
                print("Fila {}: IP address vacio".format(count))
            if ap_row[5].value != None:
                if mac_pattern.fullmatch(ap_row[5].value) != None:
                    format_valid = True
                    print("Fila {}: MAC sin errores".format(count))
                else:
                    format_valid = False
                    print("Fila {}: ERROR en MAC".format(count))
            else:
                print("Fila {}: MAC vacio".format(count))
            if ap_row[6].value != None:
                if serial_pattern.fullmatch(ap_row[6].value) != None:
                    format_valid = True
                    print("Fila {}: Serial sin errores".format(count))
                else:
                    format_valid = False
                    print("Fila {}: ERROR en Serial".format(count))
            else:
                print("Fila {}: Serial vacio".format(count))
        return format_valid

    def get_ap_operational(client_id,vendor):
        """
        Lee en la controladora los aps existentes en un cliente y retorna una lista con la información de los APs
        :param client_id: ID del cliente comprensible por la controladora
        :param vendor: vendor de los APs (Meraki, Ruckus_vsz)
        :return: Lista con elementos de la clase AccessPoint conteniendo la información leída de los APs
        """
        read_operational_aps = []
        if vendor == 'ruckus_vsz':
            if controller.sz_url and controller.login_pass:
                vSZ_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                param_st = {'serviceTicket': vSZ_comm.serviceTicket}
                body_req = {
                    "filters": [
                        {
                            "type": "DOMAIN",
                            "value": client_id
                        }
                    ],
                     "fullTextSearch": {
                         "type": "AND",
                         "value": ""
                     },
                    "attributes": [
                        "*"
                    ],
                    "limit": 999
                }
                res_operational_aps = vSZ_comm.session.post('{vSZurl}{ver}/query/ap'.format(vSZurl=controller.sz_url, ver=controller.sz_version),
                                                            params=param_st, json=body_req)
                status_aps_list = res_operational_aps.json()['list'] # Lista con información de los APs operativos en el dominio
                for ap_operational in status_aps_list:
                    ap = AccessPoint('ruckus_vsz')
                    ap.model = ap_operational['model']
                    ap.name = ap_operational['deviceName']
                    ap.description = ap_operational['description']
                    ap.mac = ap_operational['apMac']
                    ap.serial = ap_operational['serial']
                    ap.ip = ap_operational['ip']
                    ap.site = ap_operational['location']
                    ap.status = ap_operational['status']
                    ap.clients = ap_operational['numClients']
                    ap.controller_ip = controller.ip
                    read_operational_aps.append(ap)
                print("Lee los APs del cliente {}".format(client_id))
                vSZ_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("user or password not defined")
        elif vendor == 'ruckus_sz_onsite':
            if controller.sz_url and controller.login_pass:
                sz_comm = SZ_Communication(controller.sz_url, controller.sz_version, controller.login_pass, controller.login_user)
                param_st = {'serviceTicket': sz_comm.serviceTicket}
                body_req = {
                    "fullTextSearch": {
                        "type": "AND",
                        "value": ""
                    },
                    "attributes": [
                        "*"
                    ],
                    "limit": 999
                }
                res_operational_aps = sz_comm.session.post('{SZurl}{ver}/query/ap'.format(SZurl=controller.sz_url, ver=controller.sz_version),
                                                            params=param_st, json=body_req)
                status_aps_list = res_operational_aps.json()['list'] # Lista con los APs cargados en la controladora
                for ap_operational in status_aps_list:
                    ap = AccessPoint('ruckus_vsz')
                    ap.model = ap_operational['model']
                    ap.name = ap_operational['deviceName']
                    ap.description = ap_operational['description']
                    ap.mac = ap_operational['apMac']
                    ap.serial = ap_operational['serial']
                    ap.ip = ap_operational['ip']
                    ap.site = ap_operational['location']
                    ap.status = ap_operational['status']
                    ap.clients = ap_operational['numClients']
                    ap.controller_ip = controller.ip
                    read_operational_aps.append(ap)
                print("Lee los APs de la controladora")
                sz_comm.close_sz_session(controller.sz_url, controller.sz_version)
            else:
                error_window("You must define user and password")
                raise Exception("user or password not defined")
        elif vendor == 'meraki':
            meraki_comm = MerakiCommunicator(controller.meraki_api_key)
            devices = meraki_comm.send_request('https://api.meraki.com/api/v1/organizations/{}/devices'.format(client_id))
            time.sleep(0.21)
            devices_status = meraki_comm.send_request('https://api.meraki.com/api/v1/organizations/{}/devices/availabilities'.format(client_id))
            time.sleep(0.21)
            netowrks = meraki_comm.send_request('https://api.meraki.com/api/v1/organizations/{}/networks'.format(client_id))
            networks_ids = [net['id'] for net in netowrks]
            mr_pattern = re.compile('^MR\w+', re.IGNORECASE)
            for ap_operational in devices:
                if mr_pattern.fullmatch(ap_operational['model']) != None:
                    ap = AccessPoint('meraki')
                    ap.model = ap_operational['model']
                    ap.name = ap_operational['name']
                    ap.description = ap_operational['notes']
                    ap.mac = ap_operational['mac']
                    ap.serial = ap_operational['serial']
                    ap.ip = ap_operational['lanIp']
                    ap.site = ap_operational['address']
                    read_operational_aps.append(ap)
                else:
                    print("El equipo {} no es un modelo MR".format(ap_operational['serial']))
            print("Lee los APs del cliente {}".format(client_id))
            # Obtiene el listado de clientes de Meraki en todas las network de los últimos 5 minutos y crea una lista con
            # las repeticiones de seriales de APs
            client_ap_connected = []
            for netid in networks_ids:
                time.sleep(0.21)
                clients_org = meraki_comm.send_request('https://api.meraki.com/api/v1/networks/{}/clients?timespan=300'
                                                       .format(netid))
                client_ap_connected.extend([serial['recentDeviceSerial'] for serial in clients_org])
                print("Leyendo red {}".format(netid))
            # Guarda los status y cantidad de clientes de los APs
            for ap_meraki_partial in read_operational_aps:
                # Almacena el status del AP
                for indx, status in enumerate(devices_status):
                    if ap_meraki_partial.serial == status['serial']:
                        ap_meraki_partial.status = status['status']
                        devices_status.pop(indx)
                        break
                # Cuenta la cantidad de clientes del AP
                ap_meraki_partial.clients = client_ap_connected.count(ap_meraki_partial.serial)
        elif vendor == 'fortinet':

            if controller.forti_key:
                url = "https://{fIP}/api/v2/monitor/wifi/managed_ap?" \
                      "vdom=*&access_token={key}".format(fIP=controller.ip, key=controller.forti_key)
                requests.packages.urllib3.disable_warnings()
                data = requests.get(url, verify=False)
                print("request response code : .............................. ", data.status_code)
                main_dict = data.json()
                main_dic_results = []
                for vdom_ans in main_dict:
                    if len(vdom_ans['results']) > 0:
                        main_dic_results.extend(vdom_ans['results'])
                for ap_operational in range(len(main_dic_results)):
                    ap = AccessPoint('fortinet')
                    #Condicional para los casos en los que el FAP está caído y no hay key 'os_version'
                    if 'os_version' in main_dic_results[ap_operational]:
                        ap.model = main_dic_results[ap_operational]['os_version']
                    else:
                        ap.model = None  
                    print(ap.model)
                    ap.name = main_dic_results[ap_operational]['name']
                    ap.description = main_dic_results[ap_operational]['location']
                    ap.mac = main_dic_results[ap_operational]['board_mac']
                    ap.serial = main_dic_results[ap_operational]['serial']
                    ap.ip = main_dic_results[ap_operational]['local_ipv4_addr']
                    ap.status = main_dic_results[ap_operational]['status']
                    ap.clients = main_dic_results[ap_operational]['clients']
                    ap.controller_ip = controller.ip
                    read_operational_aps.append(ap)
            else:
                error_window("You must define Fortinet PI key")
                raise Exception("API key not defined")
        else:
            print("error: vendor desconocido")
        return read_operational_aps

    def fill_AP_info(vendor, new_file):
        """
        LLena un archivo excel con la información de APs de un cliente
        :param vendor: vendor de los APs (Meraki, Ruckus_vsz)
        :param new_file: Archivo excel en el que se van a escribir los valores descargados de la controladora
        """
        client_id = select_client(vendor)
        operational_aps = get_ap_operational(client_id, vendor)
        ap_op_wb = openpyxl.Workbook()
        ap_op_sh = ap_op_wb.active
        ap_op_sh['A1'] = "AP NAME"
        ap_op_sh['B1'] = "MODEL"
        ap_op_sh['C1'] = "Description/Location"
        ap_op_sh['D1'] = "Site"
        ap_op_sh['E1'] = "IP Address"
        ap_op_sh['F1'] = "MAC"
        ap_op_sh['G1'] = "Serial"
        ap_op_sh['H1'] = "Status"
        ap_op_sh['I1'] = "Clients"
        for ap in operational_aps:
            ap_op_sh.append([ap.name, ap.model, ap.description, ap.site, ap.ip, ap.mac, ap.serial, ap.status, ap.clients])
        print("APs leidos desde la controladora")
        ap_op_sh.title = 'APinfo'
        ap_op_wb.save(new_file)

    def clear_work_area():
        """
        Limpia el area de trabajo
        """
        for widgets in root_win.winfo_children():
            widgets.destroy()

    class GuiCompleteFile:
        """
        Clase con las funciones para complementar las MAC y Seriales del archivo
        """
        def __init__(self):
            self.frm = ''
            self.file_source = ''
            self.file_dest = ''
            self.vendor_selected = None
            self.login_user = ''
            self.login_pass = ''
            self.ip = ''
            self.sz_old = False
            self.current_state = tkinter.StringVar()
            self.fortikey = ''

        def select_file_original(self):
            """
            Función grafica para seleccionar el archivo con los datos iniciales a completar
            """
            self.file_source = filedialog.askopenfilename(initialdir=Path.home(), title="Select initial file",
                                                   filetypes=(("Excel", "*.xlsx"), ("All", "*.*")))
            lb_file_selected = ttk.Label(self.frm, text=self.file_source)
            lb_file_selected.grid(row=7, column=1, sticky=tkinter.W)
            self.current_state.set("Source file selected")

        def complete_file(self):
            """
            Seleccionar el archivo en el que se van a guardar los datos completados de serial y mac
            """
            self.current_state.set("Procesing...")
            self.file_dest = filedialog.asksaveasfilename(defaultextension=".xlsx", initialdir=Path.home(),
                                                         title="Select destination file",
                                                         filetypes=(('excel', '*.xlsx'), ('All', '*.*')))
            lb_file_dest = ttk.Label(self.frm, text=self.file_dest)
            lb_file_dest.grid(row=8, column=1, sticky=tkinter.W)
            print("Archivo {} seleccionado como destino".format(self.file_dest))
            controller.login_user = self.login_user.get()
            controller.login_pass = self.login_pass.get()
            controller.forti_key = self.fortikey.get()
            controller.set_ip(self.ip.get())
            controller.set_sz_api_version(self.sz_old.get())

            if self.vendor_selected:
                print(self.vendor_selected)
                complete_mac_or_serial(self.vendor_selected, self.file_source, self.file_dest)
            else:
                error_window("You need to select vendor")
                raise Exception("vendor not selected")
            self.current_state.set("Process completed\n"
                                   "please verify your destination file")

        def run_gui(self):
            """
            Ejecuta la interfaz gráfica para completar los seriales y mac
            """
            def select_vendor(event):
                vendors_label_map = {'Virtual Ruckus-vSZ': 'ruckus_vsz', 'Onsite Ruckus-SZ': 'ruckus_sz_onsite',
                                     'Meraki': 'meraki', 'Fortinet': 'fortinet'}
                self.vendor_selected = vendors_label_map[selected_vendor.get()]
                controller.vendor = self.vendor_selected
                if self.vendor_selected == 'meraki':
                    user_ent.config(state='disabled')
                    pass_ent.config(state='disabled')
                    ip_ent.config(state='disabled')
                    fortikey_ent.config(state='disabled')
                elif self.vendor_selected == 'fortinet':
                    user_ent.config(state='disabled')
                    pass_ent.config(state='disabled')
                    ip_ent.config(state='normal')
                    fortikey_ent.config(state='normal')
                else:
                    szold_chbutt.config(state='normal')
                    user_ent.config(state='normal')
                    pass_ent.config(state='normal')
                    ip_ent.config(state='normal')
                    fortikey_ent.config(state='disabled')

            clear_work_area()
            self.frm = tkinter.Frame(master=root_win)
            self.frm.grid(row=0, column=0)
            self.frm.rowconfigure(0, weight=1, minsize=10)
            self.frm.rowconfigure(1, weight=1, minsize=10)
            self.frm.rowconfigure(2, weight=1, minsize=10)
            self.frm.rowconfigure(3, weight=1, minsize=10)
            self.frm.rowconfigure(4, weight=1, minsize=10)
            self.frm.rowconfigure(5, weight=1, minsize=10)
            self.frm.rowconfigure(6, weight=1, minsize=10)
            self.frm.rowconfigure(7, weight=1, minsize=10)
            self.frm.rowconfigure(8, weight=1, minsize=10)
            self.frm.rowconfigure(9, weight=1, minsize=10)
            self.frm.columnconfigure(0, weight=1, minsize=10)
            self.frm.columnconfigure(1, weight=1, minsize=10)
            # Titulo
            title_lb = ttk.Label(self.frm, text="Complete MAC/Serial info", anchor=tkinter.CENTER, font=('Helvetica',12))
            title_lb.grid(row=0, columnspan=2, pady=10)
            # Lista desplegable de vendors
            selected_vendor = tkinter.StringVar(self.frm)
            vendor_lb = ttk.Label(self.frm, text="Choose vendor")
            vendor_lb.grid(row=1, column=0, sticky=tkinter.E)
            cb_vendors = ttk.Combobox(self.frm, state='readonly', textvariable=selected_vendor)
            cb_vendors['values'] = ['Virtual Ruckus-vSZ', 'Onsite Ruckus-SZ', 'Meraki', 'Fortinet']
            cb_vendors.grid(row=1, column=1, sticky=tkinter.W)
            cb_vendors.bind('<<ComboboxSelected>>', select_vendor)
            # Entrada para el user
            self.login_user = tkinter.StringVar(self.frm)
            user_lb = tkinter.Label(self.frm, text="Login user")
            user_lb.grid(row=2, column=0, sticky=tkinter.E)
            user_ent = tkinter.Entry(self.frm, textvariable=self.login_user, width=20)
            user_ent.grid(row=2, column=1, sticky=tkinter.W)
            # Entrada para el password
            self.login_pass = tkinter.StringVar(self.frm)
            pass_lb = tkinter.Label(self.frm, text="Password")
            pass_lb.grid(row=3, column=0, sticky=tkinter.E)
            pass_ent = tkinter.Entry(self.frm, show='*', textvariable=self.login_pass, width=20)
            pass_ent.grid(row=3, column=1, sticky=tkinter.W)
            # Entrada para la IP
            self.ip = tkinter.StringVar(self.frm)
            ip_lb = tkinter.Label(self.frm, text="Controller IP Address")
            ip_lb.grid(row=4, column=0, sticky=tkinter.E)
            ip_ent = tkinter.Entry(self.frm, textvariable=self.ip, width=20)
            ip_ent.grid(row=4, column=1, sticky=tkinter.W)
            # Checkbox smartzone version vieja
            self.sz_old = tkinter.BooleanVar()
            self.sz_old.set(False)
            szold_chbutt = ttk.Checkbutton(self.frm, text='SmartZone pre-5.2 version', variable=self.sz_old, state='disabled')
            szold_chbutt.grid(row=5, column=0)
            # Entrada para la key fortigate
            self.fortikey = tkinter.StringVar(self.frm)
            fortikey_lb = tkinter.Label(self.frm, text="Fortigate API key")
            fortikey_lb.grid(row=6, column=0, sticky=tkinter.E)
            fortikey_ent = tkinter.Entry(self.frm, textvariable=self.fortikey, show='*', width=20)
            fortikey_ent.grid(row=6, column=1, sticky=tkinter.W)
            # Boton para seleccionar archivo a completar
            btn_select_original = tkinter.Button(self.frm, text='Select initial file', command=self.select_file_original)
            btn_select_original.grid(row=7, column=0, sticky=tkinter.W)
            # Boton para generar el archivo completo
            btn_complete_file = tkinter.Button(self.frm, text='Create completed file', command=self.complete_file)
            btn_complete_file.grid(row=8, column=0, sticky=tkinter.W)
            # Mensage de estado
            state_lb = tkinter.Label(self.frm, textvariable=self.current_state)
            state_lb.grid(row=9, column=1)


    class GuiGetInfo:
        """
        Clase con la interfaz grafica para obtener la información de APs y grabarla en un excel
        """
        def __init__(self):
            self.frm = ''
            self.file_dest = ''
            self.vendor_selected = None
            self.login_user = ''
            self.login_pass = ''
            self.ip = ''
            self.sz_old = False
            self.current_state = tkinter.StringVar()
            self.fortikey = ''


        def get_ap_info(self):
            """
            Selecciona archivo excel donde se guardará la información y descarga la información del cliente
            """
            self.current_state.set("Procesing...")
            self.file_dest = filedialog.asksaveasfilename(defaultextension=".xlsx", initialdir=Path.home(),
                                                         title="Select dest file",
                                                         filetypes=(('excel', '*.xlsx'), ('All', '*.*')))
            lb_file_dest = ttk.Label(self.frm, text=self.file_dest)
            lb_file_dest.grid(row=7, column=1, sticky=tkinter.W)
            print("Archivo {} seleccionado como destino".format(self.file_dest))
            controller.login_user = self.login_user.get()
            controller.login_pass = self.login_pass.get()
            controller.forti_key = self.fortikey.get()
            controller.set_ip(self.ip.get())
            controller.set_sz_api_version(self.sz_old.get())

            if self.vendor_selected:
                print(self.vendor_selected)
                fill_AP_info(self.vendor_selected, self.file_dest)
            else:
                error_window("You need to select vendor")
                raise Exception("vendor not selected")

            self.current_state.set("Process completed\n"
                                   "please verify your destination file")

        def run_gui(self):
            """
           Ejecuta la interfaz gráfica de descargar información de APs 
            """
            def select_vendor(event):
                vendors_label_map = {'Virtual Ruckus-vSZ': 'ruckus_vsz', 'Onsite Ruckus-SZ': 'ruckus_sz_onsite', 'Meraki': 'meraki', 'Fortinet': 'fortinet'}
                self.vendor_selected = vendors_label_map[selected_vendor.get()]
                controller.vendor = self.vendor_selected
                if self.vendor_selected == 'meraki':
                    user_ent.config(state='disabled')
                    pass_ent.config(state='disabled')
                    ip_ent.config(state='disabled')
                    ip_ent.config(state='disabled')
                    fortikey_ent.config(state='disabled')
                elif self.vendor_selected == 'fortinet':
                    user_ent.config(state='disabled')
                    pass_ent.config(state='disabled')
                    ip_ent.config(state='normal')
                    fortikey_ent.config(state='normal')
                else:
                    szold_chbutt.config(state='normal')
                    user_ent.config(state='normal')
                    pass_ent.config(state='normal')
                    ip_ent.config(state='normal')
                    fortikey_ent.config(state='disabled')

            clear_work_area()
            self.frm = tkinter.Frame(master=root_win)
            self.frm.grid(row=0, column=0)
            self.frm.rowconfigure(0, weight=1, minsize=10)
            self.frm.rowconfigure(1, weight=1, minsize=10)
            self.frm.rowconfigure(2, weight=1, minsize=10)
            self.frm.rowconfigure(3, weight=1, minsize=10)
            self.frm.rowconfigure(4, weight=1, minsize=10)
            self.frm.rowconfigure(5, weight=1, minsize=10)
            self.frm.rowconfigure(6, weight=1, minsize=10)
            self.frm.rowconfigure(7, weight=1, minsize=10)
            self.frm.rowconfigure(8, weight=1, minsize=10)
            self.frm.columnconfigure(0, weight=1, minsize=10)
            self.frm.columnconfigure(1, weight=1, minsize=10)
            # Titulo
            title_lb = ttk.Label(self.frm, text="Get customer's AP info", anchor=tkinter.CENTER, font=('Helvetica',12))
            title_lb.grid(row=0, columnspan=2, pady=10)
            # Lista desplegable de vendors
            selected_vendor = tkinter.StringVar(self.frm)
            vendor_lb = ttk.Label(self.frm, text="Choose vendor")
            vendor_lb.grid(row=1, column=0, sticky=tkinter.E)
            cb_vendors = ttk.Combobox(self.frm, state='readonly', textvariable=selected_vendor)
            cb_vendors['values'] = ['Virtual Ruckus-vSZ', 'Onsite Ruckus-SZ', 'Meraki', 'Fortinet']
            cb_vendors.grid(row=1, column=1, sticky=tkinter.W)
            cb_vendors.bind('<<ComboboxSelected>>', select_vendor)
            # Entrada para el user
            self.login_user = tkinter.StringVar(self.frm)
            user_lb = tkinter.Label(self.frm, text="Login user")
            user_lb.grid(row=2, column=0, sticky=tkinter.E)
            user_ent = tkinter.Entry(self.frm, textvariable=self.login_user, width=20)
            user_ent.grid(row=2, column=1, sticky=tkinter.W)
            # Entrada para el password
            self.login_pass = tkinter.StringVar(self.frm)
            pass_lb = tkinter.Label(self.frm, text="Password")
            pass_lb.grid(row=3, column=0, sticky=tkinter.E)
            pass_ent = tkinter.Entry(self.frm, show='*', textvariable=self.login_pass, width=20)
            pass_ent.grid(row=3, column=1, sticky=tkinter.W)
            # Entrada para la IP
            self.ip = tkinter.StringVar(self.frm)
            ip_lb = tkinter.Label(self.frm, text="Controller IP Address")
            ip_lb.grid(row=4, column=0, sticky=tkinter.E)
            ip_ent = tkinter.Entry(self.frm, textvariable=self.ip, width=20)
            ip_ent.grid(row=4, column=1, sticky=tkinter.W)
            # Checkbox smartzone version vieja
            self.sz_old = tkinter.BooleanVar()
            self.sz_old.set(False)
            szold_chbutt = ttk.Checkbutton(self.frm, text='SmartZone pre-5.2 version', variable=self.sz_old, state='disabled')
            szold_chbutt.grid(row=5, column=0)
            # Entrada para la key fortigate
            self.fortikey = tkinter.StringVar(self.frm)
            fortikey_lb = tkinter.Label(self.frm, text="Fortigate API key")
            fortikey_lb.grid(row=6, column=0, sticky=tkinter.E)
            fortikey_ent = tkinter.Entry(self.frm, textvariable=self.fortikey, show='*', width=20)
            fortikey_ent.grid(row=6, column=1, sticky=tkinter.W)
            # Boton para crear archivo con informacion
            btn_get_info = tkinter.Button(self.frm, text='Get AP info', command=self.get_ap_info)
            btn_get_info.grid(row=7, column=0, sticky=tkinter.W)
            # Mensage de estado
            state_lb = tkinter.Label(self.frm, textvariable=self.current_state)
            state_lb.grid(row=8, column=1)

    class GuiLabelAps:
        """
        Clase con interfaz gráfica para marcar los APs en la controladora
        """
        def __init__(self):
            self.frm = ''
            self.file_source = ''
            self.vendor_selected = None
            self.login_user = ''
            self.login_pass = ''
            self.ip = ''
            self.sz_old = False
            self.current_state = tkinter.StringVar()

        def set_label_aps(self):
            """
            Función para seleccionar el archivo excel con origen de datos y marcar los APs en la controladora
            """
            self.current_state.set("Processing...")
            self.file_source = filedialog.askopenfilename(initialdir=Path.home(), title="Select origin file",
                                                   filetypes=(("Excel", "*.xlsx"), ("All", "*.*")))
            lb_file_selected = ttk.Label(self.frm, text=self.file_source)
            lb_file_selected.grid(row=7, column=1, sticky=tkinter.W)
            print("Archivo {} seleccionado como origen".format(self.file_source))
            controller.login_user = self.login_user.get()
            controller.login_pass = self.login_pass.get()
            controller.forti_key = self.fortikey.get()
            controller.set_ip(self.ip.get())
            controller.set_sz_api_version(self.sz_old.get())

            if self.vendor_selected:
                print(self.vendor_selected)
                label_aps(self.vendor_selected, self.file_source)
            else:
                error_window("You must select vendor")
                raise Exception("vendor not selected")

            self.current_state.set("Process completed\n"
                                   "please verify the APs in the controller")

        def run_gui(self):
            """
            Ejecuta la interfaz gráfica para marcar los APs en la controladora
            """
            def select_vendor(event):
                vendors_label_map = {'Virtual Ruckus-vSZ': 'ruckus_vsz', 'Onsite Ruckus-SZ': 'ruckus_sz_onsite', 'Meraki': 'meraki', 'Fortinet': 'fortinet'}
                self.vendor_selected = vendors_label_map[selected_vendor.get()]
                controller.vendor = self.vendor_selected
                if self.vendor_selected == 'meraki':
                    user_ent.config(state='disabled')
                    pass_ent.config(state='disabled')
                    ip_ent.config(state='disabled')
                    fortikey_ent.config(state='disabled')
                elif self.vendor_selected == 'fortinet':
                    user_ent.config(state='disabled')
                    pass_ent.config(state='disabled')
                    ip_ent.config(state='normal')
                    fortikey_ent.config(state='normal')
                else:
                    szold_chbutt.config(state='normal')
                    user_ent.config(state='normal')
                    pass_ent.config(state='normal')
                    ip_ent.config(state='normal')
                    fortikey_ent.config(state='disabled')

            clear_work_area()
            self.frm = tkinter.Frame(master=root_win)
            self.frm.grid(row=0, column=0)
            self.frm.rowconfigure(0, weight=1, minsize=10)
            self.frm.rowconfigure(1, weight=1, minsize=10)
            self.frm.rowconfigure(2, weight=1, minsize=10)
            self.frm.rowconfigure(3, weight=1, minsize=10)
            self.frm.rowconfigure(4, weight=1, minsize=10)
            self.frm.rowconfigure(5, weight=1, minsize=10)
            self.frm.rowconfigure(6, weight=1, minsize=10)
            self.frm.rowconfigure(7, weight=1, minsize=10)
            self.frm.rowconfigure(8, weight=1, minsize=10)
            self.frm.columnconfigure(0, weight=1, minsize=10)
            self.frm.columnconfigure(1, weight=1, minsize=10)
            # Titulo
            title_lb = ttk.Label(self.frm, text="Set AP name and description", anchor=tkinter.CENTER, font=('Helvetica',12))
            title_lb.grid(row=0, columnspan=2, pady=10)
            # Lista desplegable de vendors
            selected_vendor = tkinter.StringVar(self.frm)
            vendor_lb = ttk.Label(self.frm, text="Choose vendor")
            vendor_lb.grid(row=1, column=0, sticky=tkinter.E)
            cb_vendors = ttk.Combobox(self.frm, state='readonly', textvariable=selected_vendor)
            cb_vendors['values'] = ['Virtual Ruckus-vSZ', 'Onsite Ruckus-SZ', 'Meraki', 'Fortinet']
            cb_vendors.grid(row=1, column=1, sticky=tkinter.W)
            cb_vendors.bind('<<ComboboxSelected>>', select_vendor)
            # Entrada para el user
            self.login_user = tkinter.StringVar(self.frm)
            user_lb = tkinter.Label(self.frm, text="Login user")
            user_lb.grid(row=2, column=0, sticky=tkinter.E)
            user_ent = tkinter.Entry(self.frm, textvariable=self.login_user, width=20)
            user_ent.grid(row=2, column=1, sticky=tkinter.W)
            # Entrada para el password
            self.login_pass = tkinter.StringVar(self.frm)
            pass_lb = tkinter.Label(self.frm, text="Password")
            pass_lb.grid(row=3, column=0, sticky=tkinter.E)
            pass_ent = tkinter.Entry(self.frm, show='*', textvariable=self.login_pass, width=20)
            pass_ent.grid(row=3, column=1, sticky=tkinter.W)
            # Entrada para la IP
            self.ip = tkinter.StringVar(self.frm)
            ip_lb = tkinter.Label(self.frm, text="Controller IP Address")
            ip_lb.grid(row=4, column=0, sticky=tkinter.E)
            ip_ent = tkinter.Entry(self.frm, textvariable=self.ip, width=20)
            ip_ent.grid(row=4, column=1, sticky=tkinter.W)
            # Checkbox smartzone version vieja
            self.sz_old = tkinter.BooleanVar()
            self.sz_old.set(False)
            szold_chbutt = ttk.Checkbutton(self.frm, text='SmartZone pre-5.2 version', variable=self.sz_old,
                                           state='disabled')
            szold_chbutt.grid(row=5, column=0)
            # Entrada para la key fortigate
            self.fortikey = tkinter.StringVar(self.frm)
            fortikey_lb = tkinter.Label(self.frm, text="Fortigate API key")
            fortikey_lb.grid(row=6, column=0, sticky=tkinter.E)
            fortikey_ent = tkinter.Entry(self.frm, textvariable=self.fortikey, show='*', width=20)
            fortikey_ent.grid(row=6, column=1, sticky=tkinter.W)
            # Boton para configurar nombres de APs
            btn_get_info = tkinter.Button(self.frm, text='Configure APs label', command=self.set_label_aps)
            btn_get_info.grid(row=7, column=0, sticky=tkinter.W)
            # Mensage de estado
            state_lb = tkinter.Label(self.frm, textvariable=self.current_state)
            state_lb.grid(row=8, column=1)

    initial_win = tkinter.Frame(master=root_win)
    initial_win.grid(row=0, column=0, sticky="new")
    # Boton para completar la información del archivo
    gui_complete = GuiCompleteFile()
    btn_funcion1 = tkinter.Button(initial_win, text='Complete MAC/Serial info', command=gui_complete.run_gui)
    # Boton para descargar la información de los APs
    gui_ap_info = GuiGetInfo()
    btn_funcion2 = tkinter.Button(initial_win, text='Get Customer AP info', command=gui_ap_info.run_gui)
    # Boton para configurar nombres de APs
    gui_ap_labels = GuiLabelAps()
    btn_funcion3 = tkinter.Button(initial_win, text='Configure AP labels', command=gui_ap_labels.run_gui)

    btn_funcion1.grid(row=0, column=0)
    btn_funcion2.grid(row=1, column=0)
    btn_funcion3.grid(row=2, column=0)


if __name__ == '__main__':
    # fill_AP_info('ruckus_vsz')
    # complete_mac_or_serial('ruckus_vsz', config_aps_file)
    # label_aps('ruckus_vsz', config_aps_file)
    print("AP Mangement")